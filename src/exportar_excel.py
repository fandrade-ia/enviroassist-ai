"""
EnviroAssist AI — Exportación Excel
Una hoja por módulo: listado completo con columnas de categoría integradas.
"""

import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

# ── Paleta ─────────────────────────────────────────────────────────────────────
# Solo 3 colores: EP (naranja), V (amarillo), y "Otro" (verde) que engloba
# cualquier otra categoría (SAH, IE, LESRPE, "listado", etc.) — igual que en la app.
C = {
    "EP":   {"fill":"FFEDD5","font":"9A3412","label":"🟠 En peligro de extinción"},
    "V":    {"fill":"FEF9C3","font":"854D0E","label":"🟡 Vulnerable"},
    "OTRO": {"fill":"DCFCE7","font":"166534","label":"🟢 Otro"},
    "—":    {"fill":"FFFFFF","font":"374151","label":"—"},
}
HDR_FILL  = "1D4ED8"   # azul cabecera fauna
HDR_FILL_FL = "166534" # verde cabecera flora
GREY_FILL = "F3F4F6"

def _fill(h): return PatternFill(start_color=h, end_color=h, fill_type="solid")
def _font(h, bold=False, sz=10, italic=False):
    return Font(color=h, bold=bold, size=sz, name="Calibri", italic=italic)
def _bdr():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)
def _aln(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def _grupo_color(cod: str) -> str | None:
    """Agrupa cualquier código de categoría (autonómico o LESRPE, ya normalizado
    a EP/SAH/V/IE o al texto LESRPE) en los 3 colores: EP, V u OTRO."""
    if not cod or cod == "—":
        return None
    if cod == "EP":
        return "EP"
    if cod == "V":
        return "V"
    return "OTRO"


def _cat_color(cat_auto_cod, cat_lesrpe):
    """Devuelve el color más restrictivo entre autonómica y LESRPE, agrupado
    en los 3 colores del semáforo (EP > V > Otro)."""
    map_lesrpe = {
        "En peligro de extinción":"EP",
        "Vulnerable":"V",
        "LESRPE":"OTRO",
    }
    orden = {"EP":1,"V":2,"OTRO":3}
    g_a = _grupo_color(cat_auto_cod)
    g_l = _grupo_color(map_lesrpe.get(cat_lesrpe))
    candidatos = [g for g in (g_a, g_l) if g]
    if not candidatos:
        return C["—"]
    mejor = min(candidatos, key=lambda g: orden[g])
    return C[mejor]

def _hdr_row(ws, row, cols, fill_hex, font_color="FFFFFF", height=20):
    for j, text in enumerate(cols, 1):
        c = ws.cell(row=row, column=j, value=text)
        c.fill = _fill(fill_hex)
        c.font = _font(font_color, bold=True, sz=10)
        c.alignment = _aln("center")
        c.border = _bdr()
    ws.row_dimensions[row].height = height

def _leyenda(ws, row, hdr_col="1E293B"):
    """Fila de leyenda de categorías."""
    ws.cell(row=row, column=1, value="CATEGORÍAS:").font = _font(hdr_col, bold=True, sz=9)
    for j,cod in enumerate(["EP","V","OTRO"], 2):
        info = C[cod]
        c = ws.cell(row=row, column=j, value=info["label"])
        c.fill = _fill(info["fill"])
        c.font = _font(info["font"], bold=True, sz=9)
        c.alignment = _aln("center", wrap=False)
        c.border = _bdr()
    ws.row_dimensions[row].height = 16


def _hoja_inventario(ws, titulo, fill_hdr, especies_rows, alertas_idx, ccaa):
    """
    Crea hoja con listado completo + columnas de categoría integradas.
    especies_rows: list of (nombre_cientifico, nombre_comun, grupo, ...)
    alertas_idx: dict {especie: alerta_dict}
    """
    # Anchos de columna
    anchos = [32, 22, 18, 16, 22, 38]
    for j, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # Título
    ws.merge_cells("A1:F1")
    ws["A1"] = titulo
    ws["A1"].font = _font("FFFFFF", bold=True, sz=14)
    ws["A1"].fill = _fill(fill_hdr)
    ws["A1"].alignment = _aln("center")
    ws.row_dimensions[1].height = 28

    # Leyenda
    _leyenda(ws, 2)

    # Cabecera tabla
    cols = ["Nombre científico","Nombre común","Grupo",
            "LESRPE/CEEA", f"Cat. {ccaa}", "Normativa autonómica"]
    _hdr_row(ws, 3, cols, "374151", height=18)
    ws.freeze_panes = "A4"

    if not especies_rows:
        ws.merge_cells("A4:F4")
        ws["A4"] = "Sin datos"
        return

    # Filas de datos
    for i, (nombre, nombre_comun, grupo) in enumerate(especies_rows, start=4):
        a = alertas_idx.get(nombre, {})
        cat_lesrpe  = a.get("cat_lesrpe", "—")
        cat_auto_cod= a.get("cat_auto_cod", "—")
        cat_auto_nom= a.get("cat_auto_nombre", "—")
        emoji       = a.get("emoji_auto", "")
        decreto     = a.get("decreto_auto", "—")

        color = _cat_color(cat_auto_cod, cat_lesrpe)

        vals = [nombre, nombre_comun or "—", grupo or "—",
                cat_lesrpe,
                cat_auto_nom if cat_auto_cod != "—" else "—",
                decreto if decreto != "—" else "—"]

        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.fill   = _fill(color["fill"])
            c.font   = _font(color["font"], italic=(j==1), sz=10)
            c.border = _bdr()
            c.alignment = _aln("left" if j in (1,2,6) else "center")
        ws.row_dimensions[i].height = 15

    # Totales al final — desglosados por fuente normativa (LESRPE vs. autonómica)
    map_lesrpe = {"En peligro de extinción":"EP","Vulnerable":"V","LESRPE":"OTRO"}
    last = len(especies_rows) + 5
    ws.cell(row=last, column=1, value="TOTALES POR CATEGORÍA").font = _font("374151", bold=True)
    last += 1

    # Cabecera del mini-resumen: Categoría | LESRPE/CEEA | Cat. {ccaa}
    _hdr_row(ws, last, ["Categoría", "LESRPE/CEEA", f"Cat. {ccaa}"], "374151", height=16)
    last += 1

    totales_lesrpe = {}      # nº de especies según LESRPE/CEEA
    totales_auto = {}        # nº de especies según catálogo autonómico
    for a in alertas_idx.values():
        cod_a = a.get("cat_auto_cod", "—")
        cod_l = a.get("cat_lesrpe", "—")

        grupo_l = _grupo_color(map_lesrpe.get(cod_l))
        if grupo_l:
            totales_lesrpe[grupo_l] = totales_lesrpe.get(grupo_l, 0) + 1

        grupo_a = _grupo_color(cod_a)
        if grupo_a:
            totales_auto[grupo_a] = totales_auto.get(grupo_a, 0) + 1

    for cod in ["EP", "V", "OTRO"]:
        n_l = totales_lesrpe.get(cod, 0)
        n_a = totales_auto.get(cod, 0)
        if n_l or n_a:
            c1 = ws.cell(row=last, column=1, value=C[cod]["label"])
            c1.fill = _fill(C[cod]["fill"])
            c1.font = _font(C[cod]["font"], bold=True, sz=10)
            c1.border = _bdr()

            c2 = ws.cell(row=last, column=2, value=n_l)
            c2.font = _font(C[cod]["font"], sz=10)
            c2.alignment = _aln("center")
            c2.border = _bdr()

            c3 = ws.cell(row=last, column=3, value=n_a)
            c3.font = _font(C[cod]["font"], sz=10)
            c3.alignment = _aln("center")
            c3.border = _bdr()

            last += 1

    last += 1
    nota = ws.cell(row=last, column=1,
                    value="Nota: 'LESRPE/CEEA' y 'Cat. {}' cuentan especies por fuente normativa de forma "
                          "independiente; una misma especie puede figurar en ambas columnas si tiene "
                          "protección tanto nacional como autonómica.".format(ccaa))
    ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=6)
    nota.font = _font("6B7280", italic=True, sz=8)


def _hoja_graficos(ws, alertas_fauna, alertas_flora, stats_fauna, ccaa):
    """Hoja de gráficos resumen."""
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws["A1"] = "📊 Análisis Gráfico"
    ws["A1"].font = _font("1D4ED8", bold=True, sz=14)
    ws.row_dimensions[1].height = 26

    # Tabla grupos fauna
    row = 3
    ws.cell(row=row, column=1, value="Fauna por grupo taxonómico").font = _font("374151", bold=True, sz=11)
    row += 1
    _hdr_row(ws, row, ["Grupo","Nº especies"], "374151")
    row += 1
    grupos = stats_fauna.get("por_grupo", {})
    g_sorted = sorted(grupos.items(), key=lambda x: x[1], reverse=True)
    g_start = row
    for grupo, n in g_sorted:
        ws.cell(row=row, column=1, value=grupo)
        ws.cell(row=row, column=2, value=n).alignment = _aln("center")
        row += 1
    if g_sorted:
        chart = BarChart()
        chart.type = "bar"; chart.title = "Fauna por grupo"
        chart.style = 10; chart.width = 18; chart.height = 12
        chart.add_data(Reference(ws, min_col=2, min_row=g_start-1, max_row=row-1), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=g_start, max_row=row-1))
        ws.add_chart(chart, "D3")

    # Tabla categorías fauna
    row += 2
    ws.cell(row=row, column=1, value=f"Fauna protegida — {ccaa}").font = _font("374151", bold=True, sz=11)
    row += 1
    _hdr_row(ws, row, ["Categoría","Nº especies"], "374151")
    row += 1
    cat_start = row
    for cod in ["EP","V","OTRO"]:
        n = len([a for a in alertas_fauna if _grupo_color(a.get("cat_auto_cod","")) == cod])
        if n:
            c1 = ws.cell(row=row, column=1, value=C[cod]["label"])
            c1.fill = _fill(C[cod]["fill"])
            c1.font = _font(C[cod]["font"], sz=10)
            c2 = ws.cell(row=row, column=2, value=n)
            c2.alignment = _aln("center")
            row += 1
    if row > cat_start:
        pie = PieChart()
        pie.title = f"Fauna {ccaa}"; pie.style = 10
        pie.width = 14; pie.height = 10
        pie.add_data(Reference(ws, min_col=2, min_row=cat_start-1, max_row=row-1), titles_from_data=True)
        pie.set_categories(Reference(ws, min_col=1, min_row=cat_start, max_row=row-1))
        ws.add_chart(pie, "D20")


def _portada(ws, proyecto, ubicacion, ccaa, tipo, cuadriculas,
             n_fauna, n_fauna_prot, n_flora, n_flora_prot, decreto_auto):
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 52

    ws.merge_cells("A1:B1")
    ws["A1"] = "🌿 EnviroAssist AI — Informe de Inventario Biótico"
    ws["A1"].font = _font("FFFFFF", bold=True, sz=15)
    ws["A1"].fill = _fill("166534")
    ws["A1"].alignment = _aln("center")
    ws.row_dimensions[1].height = 32

    datos = [
        ("Proyecto", proyecto or "—"),
        ("Ubicación", ubicacion or "—"),
        ("Comunidad autónoma", ccaa),
        ("Tipo de proyecto", tipo or "—"),
        ("Cuadrículas IEET", ", ".join(cuadriculas) if cuadriculas else "—"),
        ("Normativa nacional", "RD 139/2011 — LESRPE/CEEA (MITECO, junio 2025)"),
        ("Normativa autonómica", decreto_auto or "No disponible"),
        ("Fecha", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Generado por", "EnviroAssist AI v1.0"),
    ]
    for i,(k,v) in enumerate(datos, 2):
        ws.cell(row=i, column=1, value=k).font = _font("374151", bold=True)
        ws.cell(row=i, column=2, value=v).font = _font("374151")
        ws.cell(row=i, column=1).fill = _fill(GREY_FILL)
        ws.cell(row=i, column=1).border = _bdr()
        ws.cell(row=i, column=2).border = _bdr()

    r = len(datos) + 3
    ws.merge_cells(f"A{r}:B{r}")
    ws[f"A{r}"] = "RESUMEN"
    ws[f"A{r}"].font = _font("FFFFFF", bold=True, sz=11)
    ws[f"A{r}"].fill = _fill(HDR_FILL)
    ws[f"A{r}"].alignment = _aln("center")

    resumen = [
        ("Total especies fauna", n_fauna),
        ("Fauna con protección", n_fauna_prot),
        ("Total taxones flora", n_flora),
        ("Flora con protección", n_flora_prot),
    ]
    for j,(k,v) in enumerate(resumen, r+1):
        ws.cell(row=j, column=1, value=k).font = _font("374151")
        ws.cell(row=j, column=2, value=v).font = _font(HDR_FILL, bold=True)
        ws.cell(row=j, column=1).border = _bdr()
        ws.cell(row=j, column=2).border = _bdr()
        ws.cell(row=j, column=2).alignment = _aln("center")

    av = r + len(resumen) + 2
    ws.merge_cells(f"A{av}:B{av}")
    ws[f"A{av}"] = "⚠️  Borrador generado con IA. Revisar y validar por técnico ambiental cualificado."
    ws[f"A{av}"].font = _font("92400E", italic=True, sz=9)
    ws[f"A{av}"].fill = _fill("FEF3C7")
    ws[f"A{av}"].alignment = _aln("left")


def generar_excel(proyecto, ubicacion, ccaa, tipo, cuadriculas,
                  alertas_fauna, alertas_flora,
                  listado_fauna, listado_flora,
                  stats_fauna, decreto_auto) -> bytes:

    wb = Workbook()
    wb.remove(wb.active)

    # Índice de alertas
    idx_fauna = {a["especie"]: a for a in alertas_fauna}
    idx_flora = {a["especie"]: a for a in alertas_flora}

    # Portada
    ws_port = wb.create_sheet("📋 Portada")
    _portada(ws_port, proyecto, ubicacion, ccaa, tipo, cuadriculas,
             len(listado_fauna) if listado_fauna is not None else 0,
             len(alertas_fauna),
             len(listado_flora),
             len(alertas_flora),
             decreto_auto)

    # Hoja fauna: listado completo con categorías
    ws_f = wb.create_sheet("🦅 Inventario Fauna")
    if listado_fauna is not None and len(listado_fauna) > 0:
        cols_ok = [c for c in ["nombre_cientifico","grupo"] if c in listado_fauna.columns]
        rows_f = []
        for _, row in listado_fauna[cols_ok].iterrows():
            nombre = row.get("nombre_cientifico","") if "nombre_cientifico" in row else row.iloc[0]
            grupo  = row.get("grupo","—") if "grupo" in row else "—"
            # Nombre común desde alertas si existe
            nc = idx_fauna.get(nombre, {}).get("nombre_comun","")
            rows_f.append((nombre, nc, grupo))
        _hoja_inventario(ws_f, f"Inventario Faunístico — {proyecto or 'Proyecto'}",
                         HDR_FILL, rows_f, idx_fauna, ccaa)
    else:
        ws_f["A1"] = "Sin datos de fauna"

    # Hoja flora: listado completo con categorías
    ws_fl = wb.create_sheet("🌿 Inventario Flora")
    if listado_flora:
        rows_fl = [(t, idx_flora.get(t,{}).get("nombre_comun",""), "Flora")
                   for t in listado_flora]
        _hoja_inventario(ws_fl, f"Inventario Florístico — {proyecto or 'Proyecto'}",
                         HDR_FILL_FL, rows_fl, idx_flora, ccaa)
    else:
        ws_fl["A1"] = "Sin datos de flora"

    # Gráficos
    ws_g = wb.create_sheet("📊 Gráficos")
    _hoja_graficos(ws_g, alertas_fauna, alertas_flora, stats_fauna, ccaa)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
